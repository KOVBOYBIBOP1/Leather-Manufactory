# -*- coding: utf-8 -*-
import openpyxl
import shutil

# Создаём копию файла
src = 'Экономика магазина кожи (1).xlsx'
dst = 'Экономика магазина кожи (1)_with_formulas.xlsx'
shutil.copy(src, dst)
print("Копия создана")

# Загружаем файл для записи формул (без data_only)
wb = openpyxl.load_workbook(dst)
ws = wb['Расход материалов']

# Очищаем текущие данные (кроме заголовков)
for row in range(2, 100):
    for col in range(1, 10):
        ws.cell(row=row, column=col).value = None
print("Данные очищены")

# Столбец A: ссылка на UNIQUE недели из Зарплатного листа
# Используем простые числа недель для начала
# Позже можно заменить на =UNIQUE('Зарплатный лист'!B2:B)

# Заголовок
ws.cell(row=1, column=1).value = "Порядок недели"

# Получим уникальные недели из файла
wb_data = openpyxl.load_workbook(src, data_only=True)
ws_econ = wb_data['Экономика магазина (полная отгр']
weeks = set()
for row in range(2, 700):
    w = ws_econ.cell(row=row, column=2).value
    if w is not None:
        try:
            weeks.add(int(w))
        except:
            pass
weeks = sorted(weeks)
wb_data.close()
print(f"Найдено {len(weeks)} недель")

# Заполняем столбец A номерами недель
for i, week in enumerate(weeks, start=2):
    ws.cell(row=i, column=1).value = week

print("Номера недель добавлены")

# Формула для расчёта расхода кожи
# Для каждой недели и каждого типа кожи суммируем:
# (количество изделий) * (расход кожи на изделие)
# где тип кожи совпадает и неделя совпадает

# Добавляем формулы для столбцов B-H
for row_idx, week in enumerate(weeks, start=2):
    for col_idx in range(2, 9):  # B до H (столбцы 2-8)
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        # Формула SUMPRODUCT
        formula = (
            f"=SUMPRODUCT("
            f"('Экономика магазина (полная отгр'!$B$2:$B$600=$A{row_idx})*"
            f"('Выполнение заказов'!$I$2:$I$600={col_letter}$1)*"
            f"('Выполнение заказов'!$P$2:$P$600)*"
            f"SUMIF('Кол-во материалов на изделие'!$A$3:$A$120,"
            f"'Выполнение заказов'!$G$2:$G$600,"
            f"'Кол-во материалов на изделие'!$B$3:$B$120))"
        )

        ws.cell(row=row_idx, column=col_idx).value = formula

    if row_idx % 10 == 0:
        print(f"Обработано {row_idx - 1} строк...")

# Сохраняем
wb.save(dst)
wb.close()
print(f"\nФайл сохранён: {dst}")
print(f"Добавлено {len(weeks)} строк с формулами для 7 типов кожи")
print("\nПример формулы в B2:")
print(f"=SUMPRODUCT(('Экономика магазина (полная отгр'!$B$2:$B$600=$A2)*('Выполнение заказов'!$I$2:$I$600=B$1)*('Выполнение заказов'!$P$2:$P$600)*SUMIF('Кол-во материалов на изделие'!$A$3:$A$120,'Выполнение заказов'!$G$2:$G$600,'Кол-во материалов на изделие'!$B$3:$B$120))")
